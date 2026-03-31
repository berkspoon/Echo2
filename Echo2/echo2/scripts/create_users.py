"""Create real user records from the active employee list + CRM data.

Step 0 of the Echo 2.0 data import. Loads ALL active employees from the
employee CSV, extracts CRM-referenced names for the name->UUID mapping,
and inserts former employees (CRM-only) as inactive users.

Usage:
    cd echo2
    python -m scripts.create_users              # dry-run (default)
    python -m scripts.create_users --apply       # actually insert
"""

import argparse
import csv
import os
import sys
import uuid
from pathlib import Path

# ---------------------------------------------------------------------------
# Setup: ensure echo2 package is importable
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from db.client import get_supabase

# ---------------------------------------------------------------------------
# Paths (relative to echo2/)
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parent.parent.parent  # Echo2/
EMPLOYEE_CSV = _ROOT / "aksia - Data Sheet - 2026-03-30.csv"
ECHO_DATA_XLSX = _ROOT / "EchoData.xlsx"
ACTIVITIES_CSV = _ROOT / "cr932_crmactivities.csv"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEV_USER_ID = "00000000-0000-0000-0000-000000000001"
SEED_EMAIL_DOMAIN = "@aksia.test"
FORMER_EMAIL_DOMAIN = "@aksia.former"
ENTRA_NAMESPACE = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

# Non-person entries in CRM data — skip these
NON_PERSON_NAMES = {"rfp team", "aksialegacy author", "rfp team "}

# Nickname aliases: CRM short form -> canonical long form on employee list
NICKNAME_MAP: dict[str, str] = {
    "tim": "timothy",
    "joe": "joseph",
    "nick": "nicholas",
    "nic": "nicholas",
    "ben": "benjamin",
    "rob": "robert",
    "tom": "thomas",
    "matt": "matthew",
    "dan": "daniel",
    "mike": "michael",
    "chris": "christopher",
    "jim": "james",
    "alex": "alexander",
    "ted": "edward",
    "bill": "william",
    "bob": "robert",
    "liz": "elizabeth",
    "beth": "elizabeth",
    "kate": "katherine",
    "katie": "katherine",
    "meg": "margaret",
    "sam": "samuel",
    "tony": "anthony",
    "rich": "richard",
    "dick": "richard",
    "steve": "stephen",
    "dave": "david",
    "jeff": "jeffrey",
    "greg": "gregory",
    "andy": "andrew",
    "pat": "patrick",
    "ed": "edward",
}

# Build reverse map for employee-has-nickname, CRM-has-full
REVERSE_NICKNAME: dict[str, list[str]] = {}
for short, full in NICKNAME_MAP.items():
    REVERSE_NICKNAME.setdefault(full, []).append(short)

# Admin email addresses
ADMIN_EMAILS = {
    "mgreenspoon@aksia.com",
    "acharalampopoulou@aksia.com",
    "asantikai@aksia.com",
    "cramsundar@aksia.com",
    "tpanopoulou@aksia.com",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _batch_insert(table_name: str, rows: list[dict], batch_size: int = 50) -> list[dict]:
    """Insert rows in batches, return all inserted records with their IDs."""
    sb = get_supabase()
    all_results: list[dict] = []
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        result = sb.table(table_name).insert(batch).execute()
        all_results.extend(result.data)
    return all_results


def normalize_name(name: str) -> str:
    """Normalize a name for matching: lowercase, strip, collapse whitespace,
    normalize hyphens to spaces."""
    name = name.lower().strip()
    name = name.replace("-", " ")
    # Collapse multiple spaces
    return " ".join(name.split())


def generate_entra_id(email: str) -> str:
    """Generate a deterministic placeholder entra_id from email."""
    return f"placeholder-{uuid.uuid5(ENTRA_NAMESPACE, email)}"


def split_name(display_name: str) -> tuple[str, str]:
    """Split a display name into (first_name, last_name)."""
    parts = display_name.strip().split()
    if len(parts) == 0:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], " ".join(parts[1:]))


def generate_former_email(display_name: str, seen_emails: set[str]) -> str:
    """Generate a placeholder email for a former employee.
    Uses @aksia.former domain to distinguish from real employees."""
    first, last = split_name(display_name)
    if first and last:
        # Remove spaces/hyphens from last name for email
        last_clean = last.replace(" ", "").replace("-", "").replace("'", "")
        base = f"{first[0].lower()}{last_clean.lower()}"
    else:
        base = display_name.lower().replace(" ", "").replace("'", "")
    email = f"{base}{FORMER_EMAIL_DOMAIN}"
    # Handle duplicates
    counter = 2
    while email in seen_emails:
        email = f"{base}{counter}{FORMER_EMAIL_DOMAIN}"
        counter += 1
    seen_emails.add(email)
    return email


# ---------------------------------------------------------------------------
# Phase 1: Load employee list
# ---------------------------------------------------------------------------

def load_employees(csv_path: Path) -> list[dict[str, str]]:
    """Load active employee list from CSV. Returns list of {name, email} dicts.
    Skips blank rows and entries with missing email."""
    employees = []
    skipped = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("Name") or "").strip()
            email = (row.get("Work email") or "").strip()
            if not name:
                continue  # blank row
            if not email:
                skipped.append(name)
                continue
            employees.append({"name": name, "email": email.lower()})
    if skipped:
        print(f"  Warning: {len(skipped)} employees with no email: {', '.join(skipped)}")
    return employees


# ---------------------------------------------------------------------------
# Phase 2: Extract CRM names
# ---------------------------------------------------------------------------

def extract_crm_names(xlsx_path: Path, csv_path: Path) -> set[str]:
    """Extract all unique person names from CRM data sources."""
    import openpyxl
    names: set[str] = set()

    # --- EchoData.xlsx Organizations sheet: coverage cols 38-42 (1-indexed) ---
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    ws_orgs = wb["Organizations"]
    org_cov_count = 0
    for row in ws_orgs.iter_rows(min_row=3, values_only=True):
        for ci in [37, 38, 39, 40, 41]:  # 0-indexed for cols 38-42
            val = row[ci] if ci < len(row) else None
            if val and isinstance(val, str) and val.strip():
                for name in val.split(";"):
                    name = name.strip().rstrip(",")
                    if name:
                        names.add(name)
                        org_cov_count += 1
    print(f"    Organizations coverage: {len(set(n.lower() for n in names))} unique names")

    # --- EchoData.xlsx Leads sheet: owner col 65 (1-indexed) ---
    ws_leads = wb["Leads"]
    lead_names_before = len(names)
    for row in ws_leads.iter_rows(min_row=2, values_only=True):
        val = row[64] if len(row) > 64 else None  # 0-indexed col 65
        if val and isinstance(val, str) and val.strip():
            for name in val.split(";"):
                name = name.strip().rstrip(",")
                if name:
                    names.add(name)
    print(f"    Lead owners: {len(set(n.lower() for n in names)) - len(set(n.lower() for n in list(names)[:lead_names_before]))} additional unique names")

    wb.close()

    # --- cr932_crmactivities.csv: cr932_author column ---
    csv.field_size_limit(sys.maxsize)
    activity_names_before = len(names)
    with open(str(csv_path), encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            author = (row.get("cr932_author") or "").strip()
            if author:
                names.add(author)
    print(f"    Activity authors: added from CSV")

    # Filter out non-person entries
    filtered = {n for n in names if normalize_name(n) not in NON_PERSON_NAMES}
    removed = len(names) - len(filtered)
    if removed:
        print(f"    Removed {removed} non-person entries")

    # Deduplicate by normalized name (keep first seen form)
    seen_normalized: dict[str, str] = {}
    for n in sorted(filtered):  # sort for determinism
        norm = normalize_name(n)
        if norm not in seen_normalized:
            seen_normalized[norm] = n

    print(f"    Total unique CRM names: {len(seen_normalized)}")
    return set(seen_normalized.keys())


# ---------------------------------------------------------------------------
# Phase 3: Name matching
# ---------------------------------------------------------------------------

def build_employee_lookup(employees: list[dict[str, str]]) -> dict[str, tuple[str, str]]:
    """Build lookup: normalized_name -> (display_name, email).
    Also indexes nickname-expanded forms."""
    lookup: dict[str, tuple[str, str]] = {}

    for emp in employees:
        name = emp["name"]
        email = emp["email"]
        norm = normalize_name(name)
        lookup[norm] = (name, email)

        # If first name has known nicknames, add those as keys too
        parts = norm.split()
        if len(parts) >= 2:
            first = parts[0]
            rest = " ".join(parts[1:])

            # If employee has long form, add short form lookups
            if first in REVERSE_NICKNAME:
                for nick in REVERSE_NICKNAME[first]:
                    nick_key = f"{nick} {rest}"
                    if nick_key not in lookup:
                        lookup[nick_key] = (name, email)

            # If employee has short form, add long form lookup
            if first in NICKNAME_MAP:
                full_key = f"{NICKNAME_MAP[first]} {rest}"
                if full_key not in lookup:
                    lookup[full_key] = (name, email)

    return lookup


def match_crm_names(
    crm_names: set[str],
    employee_lookup: dict[str, tuple[str, str]],
) -> tuple[dict[str, str], list[str], dict[str, str]]:
    """Match CRM names against employee list.

    Returns:
      - matched_emails: {email: display_name} for matched employees
      - unmatched: list of CRM display names not found in employee list
      - name_map: {normalized_crm_name: email} for ALL matched CRM names
    """
    matched_emails: dict[str, str] = {}  # email -> display_name
    unmatched: list[str] = []
    name_map: dict[str, str] = {}  # normalized_crm_name -> email
    match_details: list[tuple[str, str, str]] = []  # (crm_name, matched_to, method)

    for norm_name in sorted(crm_names):
        # Tier 1: exact normalized match
        if norm_name in employee_lookup:
            display_name, email = employee_lookup[norm_name]
            matched_emails[email] = display_name
            name_map[norm_name] = email
            match_details.append((norm_name, display_name, "exact"))
            continue

        # Tier 2: nickname expansion
        parts = norm_name.split()
        found = False
        if len(parts) >= 2:
            first = parts[0]
            rest = " ".join(parts[1:])

            # Try expanding short -> long
            if first in NICKNAME_MAP:
                expanded = f"{NICKNAME_MAP[first]} {rest}"
                if expanded in employee_lookup:
                    display_name, email = employee_lookup[expanded]
                    matched_emails[email] = display_name
                    name_map[norm_name] = email
                    match_details.append((norm_name, display_name, "nickname"))
                    found = True

            # Try contracting long -> short
            if not found and first in REVERSE_NICKNAME:
                for nick in REVERSE_NICKNAME[first]:
                    contracted = f"{nick} {rest}"
                    if contracted in employee_lookup:
                        display_name, email = employee_lookup[contracted]
                        matched_emails[email] = display_name
                        name_map[norm_name] = email
                        match_details.append((norm_name, display_name, "nickname-reverse"))
                        found = True
                        break

        if found:
            continue

        # Tier 3: prefix match (e.g. "alejandro guerra" matches "alejandro guerra a")
        for emp_key, (display_name, email) in employee_lookup.items():
            if emp_key.startswith(norm_name + " ") or norm_name.startswith(emp_key + " "):
                matched_emails[email] = display_name
                name_map[norm_name] = email
                match_details.append((norm_name, display_name, "prefix"))
                found = True
                break

        if not found:
            unmatched.append(norm_name)

    return matched_emails, unmatched, name_map, match_details


# ---------------------------------------------------------------------------
# Phase 4: Cleanup + Insert
# ---------------------------------------------------------------------------

def cleanup_existing_users(*, dry_run: bool = True) -> None:
    """Delete seed users, dev user, and former-employee placeholder users."""
    sb = get_supabase()

    if dry_run:
        # Count what would be deleted
        seed = sb.table("users").select("id", count="exact").ilike("email", f"%{SEED_EMAIL_DOMAIN}").execute()
        former = sb.table("users").select("id", count="exact").ilike("email", f"%{FORMER_EMAIL_DOMAIN}").execute()
        dev = sb.table("users").select("id").eq("id", DEV_USER_ID).execute()
        real = sb.table("users").select("id", count="exact").ilike("email", "%@aksia.com").execute()
        print(f"  [DRY-RUN] Would delete: {len(seed.data)} seed users, "
              f"{len(former.data)} former-employee users, "
              f"{'1' if dev.data else '0'} dev user, "
              f"{len(real.data)} @aksia.com users")
        return

    print("  Cleaning up existing users...")

    # First, clean seed data (records created by seed/dev users)
    from scripts.seed_data import cleanup_seed_data
    cleanup_seed_data()

    # Clean ALL tables that reference users via FK
    all_users = sb.table("users").select("id").execute().data or []
    all_user_ids = [u["id"] for u in all_users]

    # Tables that reference users — clean each one
    fk_tables = [
        ("audit_log", "changed_by"),
        ("person_coverage_owners", "user_id"),
        ("user_roles", "user_id"),
        ("lead_owners", "user_id"),
        ("documents", "uploaded_by"),
        ("saved_views", "user_id"),
    ]
    for table, col in fk_tables:
        try:
            for uid in all_user_ids:
                sb.table(table).delete().eq(col, uid).execute()
        except Exception as e:
            print(f"    Warning: {table}.{col} cleanup: {e}")

    # Also nullify created_by/coverage_owner references on entity tables
    for table, col in [
        ("organizations", "created_by"),
        ("people", "created_by"),
        ("people", "coverage_owner"),
        ("leads", "created_by"),
        ("activities", "created_by"),
        ("activities", "author_id"),
        ("contracts", "created_by"),
        ("tasks", "created_by"),
        ("tasks", "assigned_to"),
        ("distribution_lists", "created_by"),
        ("distribution_lists", "owner_id"),
    ]:
        try:
            for uid in all_user_ids:
                sb.table(table).update({col: None}).eq(col, uid).execute()
        except Exception as e:
            pass  # Column may not exist or no matching rows — fine

    # Delete ALL existing users (seed, dev, former, real — we're reimporting everything)
    for u in all_users:
        try:
            sb.table("users").delete().eq("id", u["id"]).execute()
        except Exception as e:
            print(f"    Warning: Could not delete user {u['id']}: {e}")

    print("  Cleanup complete.")


def insert_users(
    employees: list[dict[str, str]],
    former_names: list[str],
    *,
    dry_run: bool = True,
) -> tuple[list[dict], list[dict]]:
    """Insert active employees and former employees into the database.

    Returns (active_results, former_results) — lists of inserted records.
    """
    # Build active employee records
    active_records = []
    seen_emails: set[str] = set()

    for emp in employees:
        email = emp["email"]
        if email in seen_emails:
            continue  # skip duplicate emails
        seen_emails.add(email)

        first, last = split_name(emp["name"])
        role = "admin" if email in ADMIN_EMAILS else "standard_user"
        active_records.append({
            "entra_id": generate_entra_id(email),
            "email": email,
            "display_name": emp["name"],
            "first_name": first,
            "last_name": last,
            "role": role,
            "is_active": True,
        })

    # Build former employee records
    former_records = []
    former_seen_emails: set[str] = set()
    for norm_name in sorted(former_names):
        # Reconstruct a display name (title case)
        display_name = " ".join(w.capitalize() for w in norm_name.split())
        email = generate_former_email(display_name, former_seen_emails)
        first, last = split_name(display_name)
        former_records.append({
            "entra_id": generate_entra_id(email),
            "email": email,
            "display_name": display_name,
            "first_name": first,
            "last_name": last,
            "role": "standard_user",
            "is_active": False,
        })

    if dry_run:
        admin_count = sum(1 for r in active_records if r["role"] == "admin")
        print(f"\n  [DRY-RUN] Would insert:")
        print(f"    {len(active_records)} active employees ({admin_count} admins)")
        print(f"    {len(former_records)} former employees (is_active=False)")
        print(f"    {len(active_records) + len(former_records)} total users")
        if admin_count:
            print(f"\n  Admins:")
            for r in active_records:
                if r["role"] == "admin":
                    print(f"    {r['display_name']} ({r['email']})")
        return [], []

    print(f"\n  Inserting {len(active_records)} active employees...")
    active_results = _batch_insert("users", active_records)
    print(f"    -> {len(active_results)} inserted")

    print(f"  Inserting {len(former_records)} former employees (inactive)...")
    former_results = _batch_insert("users", former_records)
    print(f"    -> {len(former_results)} inserted")

    return active_results, former_results


# ---------------------------------------------------------------------------
# Phase 5: Build name->UUID mapping
# ---------------------------------------------------------------------------

def build_name_to_uuid_map() -> dict[str, str]:
    """Query the database and return a name->UUID mapping for all users.

    Returns: {normalized_display_name: user_uuid_str}
    This is the function called by downstream import steps (Steps 4-6).
    """
    sb = get_supabase()
    users = sb.table("users").select("id, display_name, email").execute().data or []

    name_map: dict[str, str] = {}
    # Build a lookup from normalized name to UUID
    norm_to_uid: dict[str, str] = {}
    for u in users:
        uid = str(u["id"])
        norm = normalize_name(u["display_name"])
        name_map[norm] = uid
        norm_to_uid[norm] = uid

        # Also map by nickname variants
        parts = norm.split()
        if len(parts) >= 2:
            first = parts[0]
            rest = " ".join(parts[1:])
            # Add short-form nicknames
            if first in REVERSE_NICKNAME:
                for nick in REVERSE_NICKNAME[first]:
                    name_map.setdefault(f"{nick} {rest}", uid)
            # Add long-form expansion
            if first in NICKNAME_MAP:
                name_map.setdefault(f"{NICKNAME_MAP[first]} {rest}", uid)

    # Add prefix-match variants: if "alejandro guerra a" is in the map,
    # also add "alejandro guerra" as a key pointing to the same UUID
    for norm, uid in list(norm_to_uid.items()):
        parts = norm.split()
        # Generate prefix variants (drop last word)
        if len(parts) >= 3:
            prefix = " ".join(parts[:-1])
            name_map.setdefault(prefix, uid)

    return name_map


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def create_users(*, dry_run: bool = True) -> dict[str, str]:
    """Main entry point.

    Returns name_to_uuid mapping after insert.
    In dry-run mode, returns empty dict.
    """
    print("\n" + "=" * 60)
    print("Echo 2.0 — Step 0: Create Users")
    print("=" * 60)
    mode = "DRY-RUN" if dry_run else "APPLY"
    print(f"Mode: {mode}\n")

    # Phase 1: Load employees
    print("Phase 1: Loading employee list...")
    employees = load_employees(EMPLOYEE_CSV)
    print(f"  {len(employees)} active employees loaded\n")

    # Phase 2: Extract CRM names
    print("Phase 2: Extracting CRM names...")
    crm_names = extract_crm_names(ECHO_DATA_XLSX, ACTIVITIES_CSV)
    print()

    # Phase 3: Match CRM names to employees
    print("Phase 3: Matching CRM names to employees...")
    employee_lookup = build_employee_lookup(employees)
    matched_emails, unmatched, name_map, match_details = match_crm_names(crm_names, employee_lookup)

    # Print matching summary
    exact = sum(1 for _, _, m in match_details if m == "exact")
    nickname = sum(1 for _, _, m in match_details if m in ("nickname", "nickname-reverse"))
    prefix = sum(1 for _, _, m in match_details if m == "prefix")
    print(f"  Exact matches: {exact}")
    print(f"  Nickname matches: {nickname}")
    print(f"  Prefix matches: {prefix}")
    print(f"  Total matched: {len(match_details)} ({len(matched_emails)} unique employees)")
    print(f"  Unmatched (former employees): {len(unmatched)}")

    if unmatched:
        print(f"\n  Former employees (not on active list):")
        for name in sorted(unmatched):
            print(f"    - {name}")

    # Print nickname match details
    nick_matches = [(c, m, t) for c, m, t in match_details if t in ("nickname", "nickname-reverse", "prefix")]
    if nick_matches:
        print(f"\n  Non-exact matches:")
        for crm_name, matched_to, method in nick_matches:
            print(f"    '{crm_name}' -> '{matched_to}' [{method}]")

    # Phase 4: Cleanup + Insert
    print(f"\nPhase 4: {'[DRY-RUN] ' if dry_run else ''}Cleanup + Insert...")
    cleanup_existing_users(dry_run=dry_run)
    active_results, former_results = insert_users(employees, unmatched, dry_run=dry_run)

    if dry_run:
        print(f"\n  [DRY-RUN] Name->UUID mapping would have {len(name_map)} CRM name entries")
        print(f"\n{'=' * 60}")
        print("Re-run with --apply to execute.")
        print("=" * 60 + "\n")
        return {}

    # Phase 5: Build name->UUID mapping from DB
    print("\nPhase 5: Building name->UUID mapping...")
    final_map = build_name_to_uuid_map()
    print(f"  {len(final_map)} name->UUID entries")

    # Verify CRM names are all in the mapping
    missing_crm = [n for n in crm_names if n not in final_map]
    if missing_crm:
        print(f"\n  WARNING: {len(missing_crm)} CRM names not in mapping:")
        for n in missing_crm:
            print(f"    - {n}")

    # Summary
    print(f"\n{'=' * 60}")
    print("Step 0 complete!")
    print(f"  Active users: {len(active_results)}")
    print(f"  Former users: {len(former_results)}")
    print(f"  Total: {len(active_results) + len(former_results)}")
    print(f"  Name->UUID entries: {len(final_map)}")
    print("=" * 60 + "\n")

    return final_map


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Create user records from employee list + CRM data"
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually insert users (default is dry-run)",
    )
    args = parser.parse_args()

    create_users(dry_run=not args.apply)
