"""Screeners router — CRUD endpoints for grid screeners (saved views).
Also provides grid inline-edit (pop-up row editor) and export endpoints.
"""

import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from db.client import get_supabase
from db.field_service import get_field_definitions, enrich_field_definitions, load_custom_values
from db.field_service import save_custom_values
from db.helpers import audit_changes
from dependencies import CurrentUser, get_current_user, require_role
from services.form_service import parse_form_data, validate_form_data, split_core_eav
from services.grid_service import (
    save_view, delete_view, set_default_view,
    update_view, duplicate_view, rename_view,
    build_grid_context,
)

router = APIRouter(prefix="/views", tags=["views"])
templates = Jinja2Templates(directory="templates")


@router.post("/save", response_class=HTMLResponse)
async def save_current_view(
    request: Request,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Save current grid state as a named view."""
    require_role(current_user, ["admin", "standard_user", "rfp_team", "legal"])

    form = await request.form()
    entity_type = form.get("entity_type", "").strip()
    view_name = form.get("view_name", "").strip()
    columns_str = form.get("columns", "")
    filters_str = form.get("filters", "{}")
    sort_by = form.get("sort_by", "").strip()
    sort_dir = form.get("sort_dir", "asc").strip()
    is_shared = form.get("is_shared") == "true"
    is_default = form.get("is_default") == "true"

    if not entity_type or not view_name:
        raise HTTPException(status_code=400, detail="Entity type and view name are required.")

    # For dashboard presets, columns is a JSON object; for grid views, comma-separated
    try:
        columns_parsed = json.loads(columns_str) if columns_str else []
    except (json.JSONDecodeError, TypeError):
        columns_parsed = None
    if isinstance(columns_parsed, (dict, list)):
        columns = columns_parsed
    else:
        columns = [c.strip() for c in columns_str.split(",") if c.strip()]
    try:
        filters = json.loads(filters_str) if filters_str else {}
    except (json.JSONDecodeError, TypeError):
        filters = {}

    # Strip internal keys from filters
    filters = {k: v for k, v in filters.items() if not k.startswith("_")}

    save_view(
        user_id=str(current_user.id),
        entity_type=entity_type,
        view_name=view_name,
        columns=columns,
        filters=filters,
        sort_by=sort_by,
        sort_dir=sort_dir,
        is_shared=is_shared,
        is_default=is_default,
    )

    # Redirect back to the entity list
    url_map = {
        "organization": "/organizations",
        "person": "/people",
        "lead": "/leads",
        "activity": "/activities",
        "contract": "/contracts",
        "task": "/tasks/my-tasks",
        "distribution_list": "/distribution-lists",
        "dashboard_advisory": "/dashboards/advisory-pipeline",
        "dashboard_capital_raise": "/dashboards/capital-raise",
    }
    redirect_url = url_map.get(entity_type, "/")
    return RedirectResponse(url=redirect_url, status_code=303)


@router.post("/{view_id}/delete", response_class=HTMLResponse)
async def delete_saved_view(
    request: Request,
    view_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Delete a saved view."""
    is_admin = current_user.role == "admin"
    success = delete_view(view_id, str(current_user.id), is_admin=is_admin)
    if not success:
        raise HTTPException(status_code=404, detail="View not found or access denied.")

    referer = request.headers.get("referer", "/")
    return RedirectResponse(url=referer, status_code=303)


@router.post("/{view_id}/set-default", response_class=HTMLResponse)
async def set_view_as_default(
    request: Request,
    view_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Mark a saved view as default for the current user."""
    form = await request.form()
    entity_type = form.get("entity_type", "").strip()
    if not entity_type:
        raise HTTPException(status_code=400, detail="Entity type is required.")

    set_default_view(view_id, str(current_user.id), entity_type)

    referer = request.headers.get("referer", "/")
    return RedirectResponse(url=referer, status_code=303)


@router.post("/{view_id}/overwrite", response_class=HTMLResponse)
async def overwrite_screener(
    request: Request,
    view_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Overwrite an existing screener with current grid state."""
    require_role(current_user, ["admin", "standard_user", "rfp_team", "legal"])

    form = await request.form()
    columns_str = form.get("columns", "")
    filters_str = form.get("filters", "{}")
    sort_by = form.get("sort_by", "").strip()
    sort_dir = form.get("sort_dir", "asc").strip()

    columns = [c.strip() for c in columns_str.split(",") if c.strip()]
    try:
        filters = json.loads(filters_str) if filters_str else {}
    except (json.JSONDecodeError, TypeError):
        filters = {}

    # Strip internal keys from filters
    filters = {k: v for k, v in filters.items() if not k.startswith("_")}

    success = update_view(
        view_id=view_id,
        user_id=str(current_user.id),
        columns=columns,
        filters=filters,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    if not success:
        raise HTTPException(status_code=404, detail="Screener not found or access denied.")

    referer = request.headers.get("referer", "/")
    return RedirectResponse(url=referer, status_code=303)


@router.post("/{view_id}/duplicate", response_class=HTMLResponse)
async def duplicate_screener(
    request: Request,
    view_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Duplicate a screener with a new name."""
    require_role(current_user, ["admin", "standard_user", "rfp_team", "legal"])

    form = await request.form()
    new_name = form.get("new_name", "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="New name is required.")

    result = duplicate_view(
        view_id=view_id,
        user_id=str(current_user.id),
        new_name=new_name,
    )
    if result is None:
        raise HTTPException(status_code=404, detail="Screener not found.")

    referer = request.headers.get("referer", "/")
    return RedirectResponse(url=referer, status_code=303)


@router.post("/{view_id}/rename", response_class=HTMLResponse)
async def rename_screener(
    request: Request,
    view_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Rename a screener."""
    require_role(current_user, ["admin", "standard_user", "rfp_team", "legal"])

    form = await request.form()
    new_name = form.get("new_name", "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="New name is required.")

    success = rename_view(
        view_id=view_id,
        user_id=str(current_user.id),
        new_name=new_name,
    )
    if not success:
        raise HTTPException(status_code=404, detail="Screener not found or access denied.")

    referer = request.headers.get("referer", "/")
    return RedirectResponse(url=referer, status_code=303)


# ---------------------------------------------------------------------------
# GRID INLINE EDIT — GET/POST /views/grid-edit/{entity_type}/{record_id}
# ---------------------------------------------------------------------------

_GRID_EDIT_TABLES = {
    "organization": "organizations",
    "person": "people",
    "lead": "leads",
    "activity": "activities",
    "contract": "contracts",
    "task": "tasks",
}


@router.get("/grid-edit/{entity_type}/{record_id}", response_class=HTMLResponse)
async def grid_edit_form(
    request: Request,
    entity_type: str,
    record_id: str,
    visible_columns: str = Query(""),
    current_user: CurrentUser = Depends(get_current_user),
):
    """HTMX partial: compact edit form for a grid row (visible columns only)."""
    if entity_type not in _GRID_EDIT_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

    table = _GRID_EDIT_TABLES[entity_type]
    sb = get_supabase()
    resp = sb.table(table).select("*").eq("id", record_id).maybe_single().execute()
    record = resp.data if resp else None
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    # Load field defs, filter to visible columns
    field_defs = get_field_definitions(entity_type, active_only=True)
    field_defs = enrich_field_definitions(field_defs)
    vis_cols = set(visible_columns.split(",")) if visible_columns else None
    if vis_cols:
        field_defs = [fd for fd in field_defs if fd["field_name"] in vis_cols]

    # Merge EAV values
    eav = load_custom_values(entity_type, record_id)
    merged = {**record, **eav}

    # Users for lookup fields
    users_resp = sb.table("users").select("id, display_name").eq("is_active", True).order("display_name").execute()
    users = users_resp.data or []

    context = {
        "request": request,
        "entity_type": entity_type,
        "record_id": record_id,
        "record": merged,
        "field_defs": field_defs,
        "users": users,
        "user": current_user,
        "errors": [],
    }
    return templates.TemplateResponse("components/_grid_edit_modal.html", context)


@router.post("/grid-edit/{entity_type}/{record_id}", response_class=HTMLResponse)
async def grid_edit_save(
    request: Request,
    entity_type: str,
    record_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Save grid inline edit. Returns HX-Trigger to refresh grid."""
    if entity_type not in _GRID_EDIT_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

    table = _GRID_EDIT_TABLES[entity_type]
    sb = get_supabase()

    # Fetch old record
    old_resp = sb.table(table).select("*").eq("id", record_id).maybe_single().execute()
    old_record = old_resp.data if old_resp else None
    if not old_record:
        raise HTTPException(status_code=404, detail="Record not found")

    # Parse form
    field_defs = get_field_definitions(entity_type, active_only=True)
    field_defs = enrich_field_definitions(field_defs)
    form = await request.form()

    # Only parse fields that were submitted (visible columns)
    submitted_fields = set(form.keys())
    visible_defs = [fd for fd in field_defs if fd["field_name"] in submitted_fields or fd["field_type"] == "boolean"]

    data = parse_form_data(entity_type, form, visible_defs)
    errors = validate_form_data(entity_type, data, visible_defs, record=old_record)

    if errors:
        eav = load_custom_values(entity_type, record_id)
        merged = {**old_record, **data}
        users_resp = sb.table("users").select("id, display_name").eq("is_active", True).order("display_name").execute()
        context = {
            "request": request,
            "entity_type": entity_type,
            "record_id": record_id,
            "record": merged,
            "field_defs": visible_defs,
            "users": users_resp.data or [],
            "user": current_user,
            "errors": errors,
        }
        return templates.TemplateResponse("components/_grid_edit_modal.html", context)

    # Split core vs EAV
    core_data, eav_data = split_core_eav(data, visible_defs)

    # Audit + update
    audit_changes(entity_type, record_id, old_record, core_data, current_user.id)
    if core_data:
        sb.table(table).update(core_data).eq("id", record_id).execute()
    if eav_data:
        save_custom_values(entity_type, record_id, eav_data, visible_defs)

    # Return empty response with trigger to refresh grid
    response = HTMLResponse("")
    response.headers["HX-Trigger"] = "gridRefresh"
    return response


# ---------------------------------------------------------------------------
# BULK OPERATIONS — POST /views/bulk-edit/{entity_type}
#                   POST /views/bulk-delete/{entity_type}
# ---------------------------------------------------------------------------

@router.post("/bulk-edit/{entity_type}", response_class=HTMLResponse)
async def bulk_edit(
    request: Request,
    entity_type: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Bulk update a single field on multiple records."""
    require_role(current_user, ["admin", "standard_user", "rfp_team"])

    if entity_type not in _GRID_EDIT_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

    form = await request.form()
    record_ids_str = form.get("record_ids", "").strip()
    field_name = form.get("field_name", "").strip()
    field_value = form.get("field_value", "")

    if not record_ids_str or not field_name:
        raise HTTPException(status_code=400, detail="record_ids and field_name are required.")

    record_ids = [rid.strip() for rid in record_ids_str.split(",") if rid.strip()]
    if not record_ids:
        raise HTTPException(status_code=400, detail="No record IDs provided.")

    table = _GRID_EDIT_TABLES[entity_type]
    sb = get_supabase()

    # Load field definitions to determine core vs EAV
    field_defs = get_field_definitions(entity_type, active_only=True)
    field_defs = enrich_field_definitions(field_defs)
    fd_map = {fd["field_name"]: fd for fd in field_defs}
    fd = fd_map.get(field_name)

    if not fd:
        raise HTTPException(status_code=400, detail=f"Unknown field: {field_name}")

    # Coerce value based on field type
    ftype = fd.get("field_type", "text")
    coerced = field_value
    if ftype == "boolean":
        coerced = field_value.lower() in ("true", "1", "yes")
    elif ftype in ("number", "currency"):
        try:
            coerced = float(field_value) if field_value else None
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid number: {field_value}")
    elif ftype == "date" and not field_value:
        coerced = None

    is_eav = fd.get("storage_type") not in (None, "core_column")
    updated = 0

    for rid in record_ids:
        # Fetch old record for audit
        old_resp = sb.table(table).select("*").eq("id", rid).maybe_single().execute()
        old_record = old_resp.data if old_resp else None
        if not old_record:
            continue

        if is_eav:
            # EAV field — save via custom values
            eav_data = {field_name: coerced}
            # Audit: load old EAV value
            old_eav = load_custom_values(entity_type, rid)
            audit_changes(entity_type, rid, {field_name: old_eav.get(field_name)}, {field_name: coerced}, current_user.id)
            save_custom_values(entity_type, rid, eav_data, [fd])
        else:
            # Core column update
            update_data = {field_name: coerced}
            audit_changes(entity_type, rid, old_record, update_data, current_user.id)
            sb.table(table).update(update_data).eq("id", rid).execute()
        updated += 1

    return HTMLResponse(f"{updated} record(s) updated.")


@router.post("/bulk-delete/{entity_type}", response_class=HTMLResponse)
async def bulk_delete(
    request: Request,
    entity_type: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Bulk soft-delete multiple records. Admin only."""
    require_role(current_user, ["admin"])

    if entity_type not in _GRID_EDIT_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

    form = await request.form()
    record_ids_str = form.get("record_ids", "").strip()

    if not record_ids_str:
        raise HTTPException(status_code=400, detail="No record IDs provided.")

    record_ids = [rid.strip() for rid in record_ids_str.split(",") if rid.strip()]
    if not record_ids:
        raise HTTPException(status_code=400, detail="No record IDs provided.")

    table = _GRID_EDIT_TABLES[entity_type]
    sb = get_supabase()
    deleted = 0

    for rid in record_ids:
        old_resp = sb.table(table).select("*").eq("id", rid).maybe_single().execute()
        old_record = old_resp.data if old_resp else None
        if not old_record or old_record.get("is_deleted"):
            continue

        audit_changes(entity_type, rid, old_record, {"is_deleted": True}, current_user.id)
        sb.table(table).update({"is_deleted": True}).eq("id", rid).execute()
        deleted += 1

    return HTMLResponse(f"{deleted} record(s) deleted.")


# ---------------------------------------------------------------------------
# EXPORT — GET /views/export/{entity_type}
# ---------------------------------------------------------------------------

_ENTITY_LABELS = {
    "organization": "Organizations",
    "person": "People",
    "lead": "Leads",
    "activity": "Activities",
    "contract": "Contracts",
    "task": "Tasks",
    "distribution_list": "Distribution_Lists",
}


@router.get("/export/{entity_type}")
async def export_to_excel(
    request: Request,
    entity_type: str,
    record_ids: str = Query(""),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Export grid data to .xlsx. Respects current filters/columns/sort."""
    if entity_type not in _GRID_EDIT_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid entity type: {entity_type}")

    from openpyxl import Workbook
    from openpyxl.styles import Font, numbers

    # Build grid context in export mode (all rows, no page cap)
    grid_ctx = build_grid_context(
        entity_type,
        request,
        current_user,
        base_url=f"/views/export/{entity_type}",
        export_mode=True,
    )

    columns = grid_ctx["columns"]
    rows = grid_ctx["rows"]

    # If record_ids provided, filter to just those
    if record_ids:
        id_set = set(rid.strip() for rid in record_ids.split(",") if rid.strip())
        rows = [r for r in rows if str(r.get("id", "")) in id_set]

    # Build workbook
    wb = Workbook()
    ws = wb.active
    label = _ENTITY_LABELS.get(entity_type, entity_type.title())
    ws.title = label

    # Header row
    header_font = Font(bold=True)
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def.get("display_name", col_def["field_name"]))
        cell.font = header_font

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_def in enumerate(columns, start=1):
            fname = col_def["field_name"]
            ftype = col_def.get("field_type", "text")
            val = row.get(fname)

            # Resolve display values for enriched fields
            if entity_type == "lead" and fname == "organization_id":
                val = row.get("org_name", val)
            elif entity_type == "lead" and fname == "aksia_owner_id":
                val = row.get("owner_name", val)
            elif entity_type == "lead" and fname == "fund_id":
                val = row.get("fund_ticker", val)
            elif entity_type == "contract" and fname == "organization_id":
                val = row.get("org_name", val)
            elif entity_type == "activity" and fname == "author_id":
                val = row.get("author_name", val)
            elif entity_type == "task" and fname == "assigned_to":
                val = row.get("assigned_to_name", val)
            elif entity_type == "person" and fname == "first_name":
                val = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()

            # Format value by type
            if val is None:
                cell = ws.cell(row=row_idx, column=col_idx, value="")
            elif ftype == "boolean":
                cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if val else "No")
            elif ftype == "currency":
                try:
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
                    cell.number_format = '#,##0'
                except (ValueError, TypeError):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            elif ftype == "number":
                try:
                    cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
                except (ValueError, TypeError):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            elif ftype == "date":
                try:
                    date_str = str(val)[:10] if val else ""
                    if date_str:
                        cell = ws.cell(row=row_idx, column=col_idx, value=datetime.strptime(date_str, "%Y-%m-%d").date())
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value="")
                except (ValueError, TypeError):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            elif ftype == "multi_select" or ftype == "text_list":
                if isinstance(val, list):
                    cell = ws.cell(row=row_idx, column=col_idx, value=", ".join(str(v) for v in val))
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            elif ftype == "dropdown":
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val).replace("_", " ").title() if val else "")
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")

    # Auto-size columns (approximate)
    for col_idx, col_def in enumerate(columns, start=1):
        max_len = len(col_def.get("display_name", col_def["field_name"]))
        for row_idx in range(2, min(len(rows) + 2, 52)):  # sample first 50 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{label}_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
